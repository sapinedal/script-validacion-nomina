import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import logging
import time

class ExcelReportExporter:
    """
    Clase encargada de exportar la comparación de nómina en un reporte Excel
    de calidad corporativa superior, aplicando estilos premium y tableros visuales.
    """
    def __init__(self, output_path: str, logger: logging.Logger = None):
        self.output_path = output_path
        self.logger = logger or logging.getLogger(__name__)

        # --- DEFINICIÓN DEL SISTEMA DE DISEÑO (ESTILOS CORPORATIVOS) ---
        self.FONT_FAMILY = "Segoe UI"
        
        # Fuentes
        self.font_title = Font(name=self.FONT_FAMILY, size=16, bold=True, color="1B365D")
        self.font_subtitle = Font(name=self.FONT_FAMILY, size=10, italic=True, color="595959")
        self.font_section = Font(name=self.FONT_FAMILY, size=12, bold=True, color="1B365D")
        self.font_header = Font(name=self.FONT_FAMILY, size=10, bold=True, color="FFFFFF")
        self.font_data = Font(name=self.FONT_FAMILY, size=10)
        self.font_data_bold = Font(name=self.FONT_FAMILY, size=10, bold=True)
        self.font_kpi_val = Font(name=self.FONT_FAMILY, size=18, bold=True, color="1B365D")
        self.font_kpi_lbl = Font(name=self.FONT_FAMILY, size=8, bold=True, color="595959")

        # Rellenos (Fills)
        self.fill_primary = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid") # Navy
        self.fill_zebra = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid") # Gris suave
        
        # Rellenos semánticos
        self.fill_match = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")     # Verde suave
        self.fill_diff = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")      # Naranja suave
        self.fill_missing = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   # Amarillo suave
        self.fill_extra = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")     # Rojo/Rosa suave
        self.fill_kpi = PatternFill(start_color="EAECEE", end_color="EAECEE", fill_type="solid")       # Fondo KPI

        # Bordes
        thin_border_side = Side(style='thin', color='D3D3D3')
        self.border_thin = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        # Borde de totales financieros (línea delgada superior, línea doble inferior)
        self.border_total = Border(
            top=Side(style='thin', color='000000'), 
            bottom=Side(style='double', color='000000')
        )

        # Alineaciones
        self.align_center = Alignment(horizontal='center', vertical='center')
        self.align_left = Alignment(horizontal='left', vertical='center')
        self.align_right = Alignment(horizontal='right', vertical='center')
        self.align_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def style_sheet(self, ws):
        """
        Activa las líneas de cuadrícula y define la fuente general en una hoja.
        """
        ws.views.sheetView[0].showGridLines = True

    def auto_fit_columns(self, ws, max_len_padding=3):
        """
        Ajusta de forma automática el ancho de las columnas de acuerdo al largo de su contenido.
        """
        for col in ws.columns:
            max_len = 0
            for cell in col:
                # Si hay salto de línea en la celda, tomar el segmento más largo
                val_str = str(cell.value or '')
                lines = val_str.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            col_letter = get_column_letter(col[0].column)
            # Definir un ancho mínimo de 10 y máximo de 45 para evitar columnas exageradamente anchas
            ws.column_dimensions[col_letter].width = min(max(max_len + max_len_padding, 10), 45)

    def write_header(self, ws, row_idx, headers):
        """
        Escribe una cabecera de tabla estilizada en una hoja.
        """
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=h)
            cell.font = self.font_header
            cell.fill = self.fill_primary
            cell.alignment = self.align_center
            cell.border = self.border_thin
        ws.row_dimensions[row_idx].height = 26

    def generate_report(self, comp_results: dict):
        """
        Genera el libro Excel final con todos los detalles estilizados.
        """
        self.logger.info(f"Generando reporte Excel estilizado en: {self.output_path}")
        
        # Asegurar que el directorio de salida existe
        out_dir = os.path.dirname(self.output_path)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir, exist_ok=True)

        wb = openpyxl.Workbook()
        
        # Eliminar hoja por defecto
        default_sheet = wb.active
        wb.remove(default_sheet)

        # 1. PESTAÑA 1: RESUMEN (Dashboard Ejecutivo)
        self._build_summary_tab(wb, comp_results)

        # 2. PESTAÑA 2: COINCIDENCIAS (Tab Color: Verde)
        self._build_data_tab(
            wb, "COINCIDENCIAS", comp_results["coincidencias"], 
            ["IDENTIFICACION", "NOMBRE", "CODIGO_NOVEDAD", "TIPO_NOVEDAD", "FECHA", "VALOR_EXCEL", "VALOR_CSV", "DIFERENCIA"],
            ["Identificación", "Nombre del Empleado", "Cód", "Tipo de Novedad", "Fecha Novedad", "Horas Excel", "Horas CSV", "Discrepancia"],
            self.fill_match, "C6EFCE"
        )

        # 3. PESTAÑA 3: DIFERENCIAS (Tab Color: Naranja)
        self._build_data_tab(
            wb, "DIFERENCIAS", comp_results["diferencias"], 
            ["IDENTIFICACION", "NOMBRE", "CODIGO_NOVEDAD", "TIPO_NOVEDAD", "FECHA", "VALOR_EXCEL", "VALOR_CSV", "DIFERENCIA"],
            ["Identificación", "Nombre del Empleado", "Cód", "Tipo de Novedad", "Fecha Novedad", "Horas Excel", "Horas CSV", "Diferencia Horas"],
            self.fill_diff, "FFD8B1"
        )

        # 4. PESTAÑA 4: FALTANTES EN PLANO (Tab Color: Amarillo)
        self._build_data_tab(
            wb, "FALTANTES_EN_PLANO", comp_results["faltantes_en_plano"], 
            ["IDENTIFICACION", "NOMBRE", "CODIGO_NOVEDAD", "TIPO_NOVEDAD", "FECHA", "VALOR_EXCEL"],
            ["Identificación", "Nombre del Empleado", "Cód", "Tipo de Novedad", "Fecha Novedad", "Horas en Excel (Sin Plano)"],
            self.fill_missing, "FFEB9C"
        )

        # 5. PESTAÑA 5: SOBRANTES EN PLANO (Tab Color: Rojo)
        self._build_data_tab(
            wb, "SOBRANTES_EN_PLANO", comp_results["sobrantes_en_plano"], 
            ["IDENTIFICACION", "NOMBRE", "CODIGO_NOVEDAD", "TIPO_NOVEDAD", "FECHA", "VALOR_CSV"],
            ["Identificación", "Nombre del Empleado", "Cód", "Tipo de Novedad", "Fecha Novedad", "Horas en Plano (Sin Excel)"],
            self.fill_extra, "FFC7CE"
        )

        # 6. PESTAÑA 6: DUPLICADOS (Tab Color: Púrpura)
        self._build_data_tab(
            wb, "DUPLICADOS", comp_results["duplicados"], 
            ["IDENTIFICACION", "NOMBRE", "CODIGO_NOVEDAD", "TIPO_NOVEDAD", "FECHA", "VALOR", "ORIGEN", "DETALLE_DUPLICADO"],
            ["Identificación", "Nombre del Empleado", "Cód", "Tipo de Novedad", "Fecha Novedad", "Horas", "Origen Duplicado", "Descripción del Riesgo"],
            None, "E1D5E7"
        )

        # Guardar archivo con manejo de fallas si está abierto
        try:
            wb.save(self.output_path)
            self.logger.info(f"Reporte Excel corporativo guardado exitosamente en: {self.output_path}")
        except PermissionError:
            base, ext = os.path.splitext(self.output_path)
            new_path = f"{base}_{int(time.time())}{ext}"
            self.logger.warning(
                f"El archivo '{self.output_path}' está abierto o bloqueado por otra aplicación (como Microsoft Excel).\n"
                f"Guardando copia en ruta alternativa: {new_path}"
            )
            wb.save(new_path)
            self.logger.info(f"Reporte alternativo guardado exitosamente en: {new_path}")
            self.output_path = new_path

    def _build_summary_tab(self, wb, comp_results: dict):
        """
        Construye la hoja 'RESUMEN' con las tarjetas de KPI y la tabla de resumen consolidada.
        """
        ws = wb.create_sheet(title="RESUMEN")
        ws.sheet_properties.tabColor = "1B365D"
        self.style_sheet(ws)

        metrics = comp_results["metrics"]

        # --- BANNER DE TÍTULO ---
        ws.cell(row=2, column=2, value="AUDITORÍA Y CONCILIACIÓN DE NOVEDADES DE NÓMINA").font = self.font_title
        ws.cell(row=3, column=2, value="Herramienta de análisis automatizado - Período de Novedades: Mayo 2026").font = self.font_subtitle
        ws.row_dimensions[2].height = 24

        # --- TARJETAS DE KPI (Dashboard) ---
        # Card 1: Excel Total Horas
        self._create_kpi_card(ws, start_row=5, start_col=2, title="HORAS EN EXCEL OPERATIVO", value=metrics["total_hours_excel"])
        # Card 2: CSV Total Horas
        self._create_kpi_card(ws, start_row=5, start_col=5, title="HORAS EN PLANO NOMINA (CSV)", value=metrics["total_hours_csv"])
        # Card 3: Horas Conciliadas (Coincidencias)
        self._create_kpi_card(ws, start_row=5, start_col=8, title="HORAS CONCILIADAS (COINCIDEN)", value=metrics["sum_hours_matches"], is_match=True)
        # Card 4: Discrepancias totales (Faltantes + Sobrantes + Diferencias de valor en plano)
        discrepancy_sum = metrics["sum_hours_missing"] + metrics["sum_hours_extra"] + abs(metrics["sum_hours_excel_diffs"] - metrics["sum_hours_csv_diffs"])
        self._create_kpi_card(ws, start_row=5, start_col=11, title="HORAS EN DISCREPANCIA (TOTAL)", value=discrepancy_sum, is_error=True)
        
        ws.row_dimensions[5].height = 14
        ws.row_dimensions[6].height = 24

        # --- SECCIÓN: TABLA DETALLADA ---
        ws.cell(row=9, column=2, value="Resumen Analítico por Tipo de Novedad").font = self.font_section
        ws.row_dimensions[9].height = 20

        df_sum = comp_results["resumen_novedades"]
        headers = df_sum.columns.tolist()
        
        # Escribir Cabecera de la tabla de resumen
        header_row = 11
        self.write_header(ws, header_row, [""] + headers)  # Espacio en columna A vacía
        
        # Desplazar 1 columna a la derecha para dejar la columna A vacía como margen
        for col_idx, h in enumerate(headers, 2):
            cell = ws.cell(row=header_row, column=col_idx, value=h)
            cell.font = self.font_header
            cell.fill = self.fill_primary
            cell.alignment = self.align_center
            cell.border = self.border_thin

        # Escribir registros de la tabla resumen
        current_row = header_row + 1
        for idx, row in df_sum.iterrows():
            ws.row_dimensions[current_row].height = 20
            
            # Aplicar celdas
            for col_idx, col_name in enumerate(headers, 2):
                val = row[col_name]
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = self.font_data
                cell.border = self.border_thin
                
                # Cebra
                if idx % 2 != 0:
                    cell.fill = self.fill_zebra
                
                # Alineaciones y formatos
                if col_idx in [2, 7, 8, 9, 10]:  # Código, cantidades
                    cell.alignment = self.align_center
                    if col_idx >= 7:
                        # Si es cantidad mayor que 0 de errores en discrepancias, resaltar ligeramente
                        if col_idx in [8, 9, 10] and val > 0:
                            cell.font = self.font_data_bold
                elif col_idx == 3:  # Nombre novedad
                    cell.alignment = self.align_left
                else:  # Horas
                    cell.alignment = self.align_right
                    cell.number_format = '#,##0.00'
                    # Si la diferencia es distinta de cero, poner en negrita
                    if col_idx == 6 and abs(val) > 0.01:
                        cell.font = self.font_data_bold
                        
            current_row += 1

        # Fila de Totales de la Tabla
        ws.row_dimensions[current_row].height = 22
        cell_lbl = ws.cell(row=current_row, column=2, value="TOTAL")
        cell_lbl.font = self.font_data_bold
        cell_lbl.border = self.border_total
        cell_lbl.alignment = self.align_center

        ws.cell(row=current_row, column=3, value="Consolidado del Período").font = self.font_data_bold
        ws.cell(row=current_row, column=3).border = self.border_total
        ws.cell(row=current_row, column=3).alignment = self.align_left

        for col_idx in range(4, 11):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = self.font_data_bold
            cell.border = self.border_total
            
            # Sumar las columnas correspondientes
            col_letter = get_column_letter(col_idx)
            # Suma desde header_row+1 hasta current_row-1
            cell.value = f"=SUM({col_letter}{header_row+1}:{col_letter}{current_row-1})"
            
            if col_idx in [7, 8, 9, 10]:
                cell.alignment = self.align_center
                cell.number_format = '#,##0'
            else:
                cell.alignment = self.align_right
                cell.number_format = '#,##0.00'

        # Ajuste general de columnas
        self.auto_fit_columns(ws, max_len_padding=3)
        # Margen de columna A fijo
        ws.column_dimensions['A'].width = 3

    def _create_kpi_card(self, ws, start_row: int, start_col: int, title: str, value: float, is_match: bool = False, is_error: bool = False):
        """
        Crea una tarjeta visual de KPI en un rango de celdas combinadas.
        """
        # Rango: Combinar 2 filas y 2 columnas
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+2)
        ws.merge_cells(start_row=start_row+1, start_column=start_col, end_row=start_row+1, end_column=start_col+2)

        # Rellenar y bordear todas las celdas implicadas para consistencia de cuadrícula de Excel
        fill_selected = self.fill_kpi
        if is_match:
            fill_selected = self.fill_match
        elif is_error and value > 0:
            fill_selected = self.fill_extra

        for r in range(start_row, start_row+2):
            for c in range(start_col, start_col+3):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill_selected
                cell.border = self.border_thin

        # Celda de Título de KPI
        title_cell = ws.cell(row=start_row, column=start_col, value=title.upper())
        title_cell.font = self.font_kpi_lbl
        title_cell.alignment = self.align_center

        # Celda de Valor
        val_cell = ws.cell(row=start_row+1, column=start_col, value=value)
        val_cell.font = self.font_kpi_val
        val_cell.alignment = self.align_center
        val_cell.number_format = '#,##0.00'
        
        # Color semántico
        if is_match:
            val_cell.font = Font(name=self.FONT_FAMILY, size=18, bold=True, color="385723")
        elif is_error and value > 0:
            val_cell.font = Font(name=self.FONT_FAMILY, size=18, bold=True, color="78281F")

    def _build_data_tab(self, wb, title: str, df: pd.DataFrame, col_keys: list, col_headers: list, fill_highlight=None, tab_color_hex: str = None):
        """
        Construye una pestaña de datos genérica (Coincidencias, Diferencias, Faltantes, Sobrantes, Duplicados).
        Aplica formateos, zebra stripes, auto-fit y colores de pestañas.
        """
        ws = wb.create_sheet(title=title)
        if tab_color_hex:
            ws.sheet_properties.tabColor = tab_color_hex
        self.style_sheet(ws)

        # Cabecera
        header_row = 1
        self.write_header(ws, header_row, col_headers)

        if df.empty:
            # Escribir aviso de vacío
            ws.row_dimensions[2].height = 25
            cell = ws.cell(row=2, column=1, value="No se detectaron registros en esta categoría de análisis.")
            cell.font = self.font_subtitle
            cell.alignment = self.align_left
            self.auto_fit_columns(ws)
            return

        # Rellenar registros
        for idx, row in df.reset_index(drop=True).iterrows():
            row_num = idx + 2
            ws.row_dimensions[row_num].height = 20
            
            is_zebra = (idx % 2 != 0)
            
            for col_idx, key in enumerate(col_keys, 1):
                val = row[key]
                
                # Manejar conversiones y formateo en base a tipos
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = self.font_data
                cell.border = self.border_thin
                
                # Cebra o resalte por defecto
                if fill_highlight:
                    cell.fill = fill_highlight
                elif is_zebra:
                    cell.fill = self.fill_zebra

                # Escribir valor
                if pd.isnull(val):
                    cell.value = "-"
                    cell.alignment = self.align_center
                else:
                    cell.value = val
                    
                    # Formateo semántico por tipo de columna
                    if key in ["VALOR_EXCEL", "VALOR_CSV", "VALOR", "DIFERENCIA"]:
                        cell.alignment = self.align_right
                        cell.number_format = '#,##0.00'
                        # Si es la columna de diferencia, negrita
                        if key == "DIFERENCIA" and abs(val) > 0.01:
                            cell.font = self.font_data_bold
                    elif key in ["IDENTIFICACION", "CODIGO_NOVEDAD", "FECHA", "ORIGEN"]:
                        cell.alignment = self.align_center
                    elif key in ["NOMBRE", "TIPO_NOVEDAD"]:
                        cell.alignment = self.align_left
                    else:
                        cell.alignment = self.align_wrap

        # Ajuste columnas
        self.auto_fit_columns(ws, max_len_padding=3)
